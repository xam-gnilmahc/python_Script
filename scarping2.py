from supabase import create_client, Client
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.drawing.image import Image
import os
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import smtplib
import ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import re
import random



class EcommerceScraper:
    def __init__(self, product_names, image_dir='downloaded_images'):
        self.product_names = product_names
        self.base_url = "https://www.flipkart.com/search?q={}&otracker=search&otracker1=search&marketplace=FLIPKART&as-show=on&as=off"
        self.image_dir = image_dir
        self.headers = ["image", "name", "price", "product_link"]
        self.sender_email = 'maxrai788@gmail.com'
        self.sender_password = 'rqcuswodywcazihj'
        self.recipients = ["maxrai788@gmail.com", "max.c@shikhartech.com"]
        self.create_excel_workbook()
        self.setup_selenium()
        self.supabase_url = "https://fzliiwigydluhgbuvnmr.supabase.co"
        self.supabase_key = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZ6bGlpd2lneWRsdWhnYnV2bm1yIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NDE5MjkxNTMsImV4cCI6MjA1NzUwNTE1M30.w3Y7W14lmnD-gu2U4dRjqIhy7JZpV9RUmv8-1ybQ92w"
        self.supabase: Client = create_client(self.supabase_url, self.supabase_key)


    def create_excel_workbook(self):
        os.makedirs(self.image_dir, exist_ok=True)
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active
        self.sheet.append(self.headers)

    def setup_selenium(self):
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("start-maximized")
        chrome_options.add_argument("disable-infobars")
        chrome_options.add_argument("--disable-extensions")
        self.driver = webdriver.Chrome(options=chrome_options)

    def scrape(self):
        for product_name in self.product_names:
            print(f"Scraping product name: {product_name}")
            search_url = self.base_url.format(product_name.replace(" ", "%20"))
            self.scrape_url(search_url) 
            print(f"Finished scraping for product: {product_name}\n")

        self.save_to_excel()

    def scrape_url(self, url):
        self.driver.get(url)
        soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        product_containers = self.get_product_containers(soup)
        
        if product_containers:
            for container in product_containers:
                self.scrape_product(container)
        else:
            print(f"No products found for URL: {url}")

    def get_product_containers(self, soup):
        # Try multiple Flipkart container classes
        return soup.find_all('div', class_=["tUxRFH", "IRpwTa", "s1Q9rs", "_1sdMkc LFEi7Z","_4WELSP","_4WELSP"])


    def scrape_product(self, container):
        image_url, product_name = self.get_product_image(container)    
        
        if not image_url or not product_name:
            print(image_url , product_name)
            print("Skipping: Missing product name or image URL.")
            return
        
        product_link = self.get_product_link(container)
        price = self.get_product_price(container)

        image = self.download_image(image_url, product_name)
        self.add_to_sheet(product_name, price, product_link)
        self.insert_into_supabase(product_name, price, product_link,image)
        
    def get_product_image(self, container):
        img_tag = container.find('img')
        if img_tag:
            image_url = img_tag.get('src') or img_tag.get('data-src') or img_tag.get('data-image')
            product_name = img_tag.get('alt', 'Image not available')
            return image_url, product_name
        return None, "Image not available"

    def insert_into_supabase(self, name, price, link, image_url):
        try:
            # If price is missing or empty, assign random price between 300 and 1000
            if not price or not isinstance(price, str) or price.strip() == "":
                clean_price = random.randint(300, 1000)
            else:
                # Remove ₹ and other non-numeric characters
                clean_price = re.sub(r'[^\d.]', '', price)
                clean_price = float(clean_price) if clean_price else random.randint(300, 1000)

            # Generate random rating between 3.0 and 5.0
            rating = round(random.uniform(3.0, 5.0), 1)

            # Generate random reviews count between 50 and 5000
            reviews_count = random.randint(50, 5000)

            data = {
                "name": name,
                "amount": clean_price,
                "type": "bot",
                "category": "bot",
                "brand": "Amazon",
                "rating": rating,
                "reviews_count": reviews_count,
                "description": name,
                "banner_url": image_url
            }

            self.supabase.table("products").insert(data).execute()
            print(f"Inserted into Supabase: {name}")
        except Exception as e:
            print(f"Failed to insert {name} into Supabase:", e)

    def get_product_link(self, container):
        detail_link = container.find('a', class_=["CGtC98", "IRpwTa", "s1Q9rs"])
        if detail_link and 'href' in detail_link.attrs:
            return "https://www.flipkart.com" + detail_link['href']
        return 'Link not available'


    def get_product_price(self, container):
        price_tag = container.find('div', class_=["Nx9bqj _4b5DiR", "rgWa7D"])
        return price_tag.text.strip() if price_tag else 'Price not available'


    def download_image(self, image_url, product_name):
        if image_url:
            image_response = requests.get(image_url)
            if image_response.status_code == 200:
                image = self.save_image(image_response.content, product_name)
            else:
                image = False
                print(f"Failed to download image for {product_name}")
        else:
            print(f"No image URL available for {product_name}")
        return image

    def save_image(self, image_content, product_name):
        safe_product_name = re.sub(r'[\\/*?:"<>|]', "", product_name)  # Clean filename
        date_str = datetime.now().strftime("%Y%m%d")  # e.g., 20250512
        rand_num = random.randint(1000, 9999)
        filename = f"{safe_product_name}_{date_str}_{rand_num}.jpg"
        image_path = os.path.join(self.image_dir, filename)

        # Try to check if file already exists on Supabase
        try:
            existing_files = self.supabase.storage.from_('productimages').list()
            if any(f['name'] == filename for f in existing_files):
                print(f"Image already exists: {filename}, skipping upload.")
                return filename
        except Exception as e:
            print(f"Error checking existing files: {e}")

        # Save image locally and upload
        with open(image_path, 'wb') as f:
            f.write(image_content)

        with open(image_path, 'rb') as f:
            self.supabase.storage.from_('productimages').upload(
                path=filename,
                file=f,
                file_options={"content-type": "image/jpeg"}
            )

        self.add_image_to_excel(image_path)
        
        return filename



    def add_image_to_excel(self, image_filename):
        img = Image(image_filename)
        img.height = 60
        img.width = 60
        self.sheet.add_image(img, f"A{self.sheet.max_row}")

    def add_to_sheet(self, product_name, price, product_link):
        self.sheet.append(['', product_name, price, product_link])

    def connect_to_smtp_server(self):
        try:
            context = ssl.create_default_context()
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls(context=context)
            server.login(self.sender_email, self.sender_password)
            return server
        except Exception as e:
            print("Error connecting to SMTP server:", e)
            return None

    def send_email(self, file_path):
        try:
            msg = MIMEMultipart()
            msg['From'] = self.sender_email
            msg['To'] = ", ".join(self.recipients)
            msg['Subject'] = "Scraping Data"

            body = "Please find attached the scraped data file."
            msg.attach(MIMEText(body, 'plain'))

            attachment = open(file_path, "rb")
            part = MIMEBase('application', 'octet-stream')
            part.set_payload((attachment).read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', "attachment; filename= " + file_path.split('/')[-1])
            msg.attach(part)

            server = self.connect_to_smtp_server()
            if server:
                text = msg.as_string()
                server.sendmail(self.sender_email, self.recipients, text)
                server.quit()
                return True
            else:
                return False

        except Exception as e:
            print("Error sending email:", e)
            return False

    def save_to_excel(self):
        if self.sheet.max_row > 1:  
            # excel_file_path = f'D:/task/scrapping_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
            # self.wb.save(excel_file_path)
            # print(f"Excel file saved at: {excel_file_path}")

            # if self.send_email(excel_file_path):
            #     print("Email sent successfully")
            # else:
            #     print("Failed to send email")
            print("done")
            print("No data to save. Excel file not created and no email sent.")


if __name__ == "__main__":
    user_input = input("Enter the product names (comma-separated): ")
    product_names = [name.strip() for name in user_input.split(",")]

    scraper = EcommerceScraper(product_names)
    scraper.scrape()
